import io
from datetime import date
from typing import Optional

import pandas as pd
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.database.db import get_db
from app.models.expense import Expense
from app.models.user import User
from app.utils.auth import get_current_user

router = APIRouter(prefix="/reports", tags=["Reports"])


def _fetch_expenses(
    db: Session,
    user_id: int,
    start_date: Optional[date],
    end_date: Optional[date],
):
    """Fetch expenses filtered by optional date range."""
    query = db.query(Expense).filter(Expense.user_id == user_id)
    if start_date:
        query = query.filter(Expense.date >= start_date)
    if end_date:
        query = query.filter(Expense.date <= end_date)
    return query.order_by(Expense.date.desc()).all()


def _build_dataframe(expenses) -> pd.DataFrame:
    """Convert a list of Expense ORM objects to a pandas DataFrame."""
    rows = []
    for e in expenses:
        rows.append(
            {
                "Date": str(e.date),
                "Description": e.description,
                "Category": e.category.name if e.category else "",
                "Amount": e.amount,
                "Payment Method": e.payment_method,
            }
        )
    return pd.DataFrame(
        rows,
        columns=["Date", "Description", "Category", "Amount", "Payment Method"],
    )


@router.get("/csv")
def export_csv(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export all user expenses as a CSV file."""
    expenses = _fetch_expenses(db, current_user.id, start_date, end_date)
    df = _build_dataframe(expenses)

    output = io.StringIO()
    df.to_csv(output, index=False)
    output.seek(0)

    filename = f"expenses_{current_user.id}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/excel")
def export_excel(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export all user expenses as an Excel (.xlsx) file."""
    expenses = _fetch_expenses(db, current_user.id, start_date, end_date)
    df = _build_dataframe(expenses)

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # Write header row with styling
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["Date", "Description", "Category", "Amount", "Payment Method"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.cell(row=row_idx, column=1, value=row.Date)
        ws.cell(row=row_idx, column=2, value=row.Description)
        ws.cell(row=row_idx, column=3, value=row.Category)
        ws.cell(row=row_idx, column=4, value=row.Amount)
        ws.cell(row=row_idx, column=5, value=getattr(row, "Payment Method"))

    # Auto-fit column widths
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"expenses_{current_user.id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/preview")
def preview_expenses(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return a JSON preview of expenses for the reports page (no file download)."""
    expenses = _fetch_expenses(db, current_user.id, start_date, end_date)
    return [
        {
            "id": e.id,
            "date": str(e.date),
            "description": e.description,
            "category_name": e.category.name if e.category else "",
            "amount": e.amount,
            "payment_method": e.payment_method,
        }
        for e in expenses
    ]

